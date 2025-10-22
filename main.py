import os
import requests
import time
import openpyxl
import json
import hashlib
import re
from datetime import datetime

# --------------------------
# Excel 工具函数
# --------------------------
def detect_last_col(ws):
    last_col = 1
    for col in range(1, ws.max_column + 1):
        column_values = [ws.cell(row=row, column=col).value for row in range(1, ws.max_row + 1)]
        if any(v is not None and v != "" for v in column_values):
            last_col = col
    return last_col

def set_column_style(ws, col_idx):
    for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.font = openpyxl.styles.Font(name="宋体", size=12)
            cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    ws.cell(row=2, column=col_idx).font = openpyxl.styles.Font(name="宋体", size=12, bold=True)

def safe_float_convert(val):
    try:
        return round(float(val), 2)
    except Exception:
        return val

def write_number_cell(ws, row, col, value):
    if isinstance(value, (int, float)):
        rounded = round(float(value), 2)
        ws.cell(row=row, column=col, value=rounded)
        ws.cell(row=row, column=col).number_format = '0.00'
    else:
        ws.cell(row=row, column=col, value=value)

# --------------------------
# 股票与指数配置
# --------------------------
stock_index_start_row = 7
stock_val_start_row = 6

# 定义字典
stocks_index = {
  "上证点数": {
    "code": "000001",
    "row": 4,
    "result": "",
    "url": "https://qt.gtimg.cn/?q=s_sh",
  },
  "中证A500": {
    "code": "000510",
    "row": 0,
    "result": "",
    "url": "https://qt.gtimg.cn/?q=s_sh",
  },
  "沪深300": {
    "code": "000300",
    "row": 0,
    "result": "",
    "url": "https://qt.gtimg.cn/?q=s_sh",
  },
  "中证500": {
    "code": "000905",
    "row": 0,
    "result": "",
    "url": "https://qt.gtimg.cn/?q=s_sh",
  },
  "沪港深500": {
    "code": "CSIH30455",
    "row": 0,
    "result": "",
    "url": "https://stock.xueqiu.com/v5/stock/realtime/quotec.json?symbol=",
  },
  "标普500": {
    "code": "usINX",
    "row": 0,
    "result": "",
    "url": "https://qt.gtimg.cn/?q=s_",
  },
  "印度": {
    "code": "SENSEX",
    "row": 0,
    "result": "",
    "url": "https://w.sinajs.cn/list=znb_",
  },
  "德国": {
    "code": "DAX_i",
    "row": 0,
    "result": "",
    "url": "https://w.sinajs.cn/list=znb_",
  },
  "日本": {
    "code": "NKY_i",
    "row": 0,
    "result": "",
    "url": "https://w.sinajs.cn/list=znb_",
  },
  "中证红利": {
    "code": "000922",
    "row": 0,
    "result": "",
    "url": "https://qt.gtimg.cn/?q=s_sh",
  },
  "红利质量": {
    "code": "CSI931468",
    "row": 0,
    "result": "",
    "url": "https://stock.xueqiu.com/v5/stock/realtime/quotec.json?symbol=",
  },
  "创业板50": {
    "code": "399673",
    "row": 0,
    "result": "",
    "url": "https://qt.gtimg.cn/?q=s_sz",
  },
  "创业板指": {
    "code": "399006",
    "row": 0,
    "result": "",
    "url": "https://qt.gtimg.cn/?q=s_sz",
  },
  "中证医疗": {
    "code": "399989",
    "row": 0,
    "result": "",
    "url": "https://qt.gtimg.cn/?q=s_sz",
  },
  "300医药": {
    "code": "000913",
    "row": 0,
    "result": "",
    "url": "https://qt.gtimg.cn/?q=s_sh",
  },
  "消费龙头": {
    "code": "CSI931068",
    "row": 0,
    "result": "",
    "url": "https://stock.xueqiu.com/v5/stock/realtime/quotec.json?symbol=",
  },
  "家用电器": {
    "code": "CSI930697",
    "row": 0,
    "result": "",
    "url": "https://stock.xueqiu.com/v5/stock/realtime/quotec.json?symbol=",
  },
  "中证白酒": {
    "code": "399997",
    "row": 0,
    "result": "",
    "url": "https://qt.gtimg.cn/?q=s_sz",
  },
  "中证消费": {
    "code": "000932",
    "row": 0,
    "result": "",
    "url": "https://qt.gtimg.cn/?q=s_sh",
  },
  "恒生医药": {
    "code": "HKHSHKBIO",
    "row": 0,
    "result": "",
    "url": "https://stock.xueqiu.com/v5/stock/realtime/quotec.json?symbol=",
  },
  "中概互联": {
    "code": "CSIH30533",
    "row": 0,
    "result": "",
    "url": "https://stock.xueqiu.com/v5/stock/realtime/quotec.json?symbol=",
  },
  "中证中药": {
    "code": "CSI930641",
    "row": 0,
    "result": "",
    "url": "https://stock.xueqiu.com/v5/stock/realtime/quotec.json?symbol=",
  },
  "恒生互联网": {
    "code": "HKHSIII",
    "row": 0,
    "result": "",
    "url": "https://stock.xueqiu.com/v5/stock/realtime/quotec.json?symbol=",
  },
  "恒生科技": {
    "code": "HKHSTECH",
    "row": 0,
    "result": "",
    "url": "https://stock.xueqiu.com/v5/stock/realtime/quotec.json?symbol=",
  },
  "全指医药": {
    "code": "000991",
    "row": 0,
    "result": "",
    "url": "https://qt.gtimg.cn/?q=s_sh",
  },
  "保险": {
    "code": "399809",
    "row": 0,
    "result": "",
    "url": "https://qt.gtimg.cn/?q=s_sz",
  },
  "中证新能源": {
    "code": "399808",
    "row": 0,
    "result": "",
    "url": "https://qt.gtimg.cn/?q=s_sz",
  },
  "中证光伏": {
    "code": "CSI931151",
    "row": 0,
    "result": "",
    "url": "https://stock.xueqiu.com/v5/stock/realtime/quotec.json?symbol=",
  },
  "新能源车": {
    "code": "399417",
    "row": 0,
    "result": "",
    "url": "https://qt.gtimg.cn/?q=s_sz",
  },
  "CS创新药": {
    "code": "CSI931152",
    "row": 0,
    "result": "",
    "url": "https://stock.xueqiu.com/v5/stock/realtime/quotec.json?symbol=",
  },
  "医疗器械": {
    "code": "BK0044",
    "row": 0,
    "result": "",
    "url": "https://stock.xueqiu.com/v5/stock/realtime/quotec.json?symbol=",
  },
  "半导体": {
    "code": "CSIH30184",
    "row": 0,
    "result": "",
    "url": "https://stock.xueqiu.com/v5/stock/realtime/quotec.json?symbol=",
  },
  "中证军工": {
    "code": "399967",
    "row": 0,
    "result": "",
    "url": "https://qt.gtimg.cn/?q=s_sz",
  },
  "中证畜牧": {
    "code": "CSI930707",
    "row": 0,
    "result": "",
    "url": "https://stock.xueqiu.com/v5/stock/realtime/quotec.json?symbol=",
  },
  "证券行业": {
    "code": "399975",
    "row": 0,
    "result": "",
    "url": "https://qt.gtimg.cn/?q=s_sz",
  },
  "中证有色": {
    "code": "CSI930708",
    "row": 0,
    "result": "",
    "url": "https://stock.xueqiu.com/v5/stock/realtime/quotec.json?symbol=",
  },
  "基建工程": {
    "code": "399995",
    "row": 0,
    "result": "",
    "url": "https://qt.gtimg.cn/?q=s_sz",
  },
  "国证地产": {
    "code": "399393",
    "row": 0,
    "result": "",
    "url": "https://qt.gtimg.cn/?q=s_sz",
  }
}
pe_pb_xilv={
"沪深全A(万德全A)":{
    "row":3,
    "code":"881001.WI",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"中证A500":{
    "row": 0,
    "code":"000510.SH",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"沪深300":{
    "row": 0,
    "code":"000300.SH",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"中证500":{
    "row": 0,
    "code":"000905.SH",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"沪港深500":{
    "row": 0,
    "code":"H30455.CSI",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"标普500":{
    "row": 0,
    "code":"SPX.GI",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"印度(印度孟买SENSEX30)":{
    "row": 0,
    "code":"SENSEX.BO",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"德国(德国DAX)":{
    "row": 0,
    "code":"GDAXI.GI",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"日本(日经225)":{
    "row": 0,
    "code":"N225.GI",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"中证红利":{
    "row": 0,
    "code":"000922.CSI",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.3,0.3,0.4]
},
"红利质量":{
    "row": 0,
    "code":"931468.CSI",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"创业板50":{
    "row": 0,
    "code":"399673.SZ",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"创业板指":{
    "row": 0,
    "code":"399006.SZ",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"中证医疗":{
    "row": 0,
    "code":"399989.SZ",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"300医药":{
    "row": 0,
    "code":"000913.SH",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"消费龙头":{
    "row": 0,
    "code":"931068.CSI",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"家用电器":{
    "row": 0,
    "code":"930697.CSI",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"中证白酒":{
    "row": 0,
    "code":"399997.SZ",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"中证消费":{
    "row": 0,
    "code":"000932.SH",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"恒生医药(恒生医疗保健)":{
    "row": 0,
    "code":"HSHCI.HI",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"中概互联(中国互联网50)":{
    "row": 0,
    "code":"H30533.CSI",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"中证中药":{
    "row": 0,
    "code":"930641.CSI",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"恒生互联网(恒生互联网科技业)":{
    "row": 0,
    "code":"HSIII.HI",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"恒生科技(恒生科技指数)":{
    "row": 0,
    "code":"HSTECH.HI",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"全指医药":{
    "row": 0,
    "code":"000991.SH",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"保险(保险II(申万))": {
  "row": 0,
  "code":"801194.SI",
  "pe_new_value":0,
  "pe_new_percent_value":0,
  "pb_new_value":0,
  "pb_new_percent_value":0,
  "xilv_new_value":0,
  "xilv_new_percent_value":0,
  "calc":[0.5,0.5,0]
},
"中证新能源(中证新能)":{
    "row": 0,
    "code":"399808.SZ",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"中证光伏(光伏产业)":{
    "row": 0,
    "code":"931151.CSI",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"新能源车":{
    "row": 0,
    "code":"930997.CSI",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"CS创新药":{
    "row": 0,
    "code":"931152.CSI",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"医疗器械":{
    "row": 0,
    "code":"h30217.CSI",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"半导体(中证全指半导体)":{
    "row": 0,
    "code":"h30184.CSI",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.5,0.5,0]
},
"中证军工":{
    "row": 0,
    "code":"399967.SZ",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.3,0.7,0]
},
"中证畜牧":{
    "row": 0,
    "code":"930707.CSI",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.3,0.7,0]
},
"证券行业(证券公司)":{
    "row": 0,
    "code":"399975.SZ",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.3,0.7,0]
},
"中证有色":{
    "row": 0,
    "code":"930708.CSI",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.3,0.7,0]
},
"基建工程(中证基建工程)":{
    "row": 0,
    "code":"399995.SZ",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.3,0.7,0]
},
"国证地产(中证全指房地产)":{
    "row": 0,
    "code":"931775.CSI",
    "pe_new_value":0,
    "pe_new_percent_value":0,
    "pb_new_value":0,
    "pb_new_percent_value":0,
    "xilv_new_value":0,
    "xilv_new_percent_value":0,
    "calc":[0.3,0.7,0]
}
}

# 自动补充行号
row_value = stock_index_start_row
for v in stocks_index.values():
    if v["row"] == 0:
        v["row"] = row_value
        row_value += 3

row_value = stock_val_start_row
for v in pe_pb_xilv.values():
    if v["row"] == 0:
        v["row"] = row_value
        row_value += 3

# --------------------------
# 请求设置
# --------------------------
session = requests.Session()
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:145.0) Gecko/20100101 Firefox/145.0",
    "Accept": "*/*",
    "Accept-Language": "zh-CN",
    "Accept-Encoding": "gzip, deflate, br, zstd",
    "Sec-Fetch-Storage-Access": "none",
    "DNT": "1",
    "Sec-GPC": "1",
    "Connection": "keep-alive",
    "Referer": "https://finance.sina.com.cn/",
    "Sec-Fetch-Dest": "script",
    "Sec-Fetch-Mode": "no-cors",
    "Sec-Fetch-Site": "cross-site",
}

# --------------------------
# 抓取指数数据
# --------------------------
def fetch_stock_data_to_ws(ws, target_col):
    for name, data in stocks_index.items():
        url = data["url"] + data["code"]
        try:
            response = session.get(url, headers=headers, timeout=10)
            response.raise_for_status()

            # 解析数据
            if "qt.gtimg.cn" in url:
                # qt.gtimg.cn 的响应以 ~ 分隔，价格一般在第 4 个位置（索引 3）
                parts = response.text.split("~")
                if len(parts) > 3 and parts[3] != '':
                    parsed = parts[3]
                else:
                    parsed = parts[0]
                data["result"] = parsed
            elif "xueqiu.com" in url:
                # 解析 xueqiu.com 的 JSON 响应
                json_data = response.json()
                data["result"] = json_data["data"][0]["current"]  # 提取 current 值
            elif "sinajs.cn" in url:
                # 解析 sina 的响应；保守匹配浮点数
                match = re.search(r'([0-9]+\.[0-9]+)', response.text)
                if match:
                    data["result"] = match.group(1)

            print(f"{name}: {data.get('result')}")
            if data["row"] == 4:
                # 写入日期和标题
                ws.cell(row=1, column=target_col, value=datetime.now().strftime("%Y/%m/%d"))
                ws.cell(row=2, column=target_col, value="上证")
            # 尝试写入浮点并限制两位小数
            val = data.get("result", "")
            numeric_val = safe_float_convert(val)
            write_number_cell(ws, data["row"], target_col, numeric_val)
        except Exception as e:
            print(f"请求 {name} 数据失败: {e}")
        # 轻微等待，避免请求过快
        time.sleep(1)

# --------------------------
# 获取 PE / PB / Xilv 数据
# --------------------------
def split_md5(md5_string, ts, gu_code):
    return {
        "gu_code": gu_code,
        "pe_category": "pe",
        "year": -1,
        "category": "",
        "ver": "new",
        "type": "pc",
        "version": "2.2.7",
        "authtoken": "",
        "act_time": ts,
        'yi854tew': md5_string[29:31],
        'u54rg5d': md5_string[2:4],
        'bioduytlw': md5_string[5:6],
        'nkjhrew': md5_string[26:27],
        'bvytikwqjk': md5_string[6:8],
        'tiklsktr4': md5_string[1:2],
        'tirgkjfs': md5_string[0:2],
        'bgd7h8tyu54': md5_string[6:8],
        'yt447e13f': md5_string[8:9],
        'nd354uy4752': md5_string[30:31],
        'ghtoiutkmlg': md5_string[11:14],
        'y654b5fs3tr': md5_string[11:12],
        'fjlkatj': md5_string[2:5],
        'jnhf8u5231': md5_string[9:11],
        'sbnoywr': md5_string[23:25],
        'kf54ge7': md5_string[31:32],
        'hy5641d321t': md5_string[25:27],
        'bgiuytkw': md5_string[9:11],
        'quikgdky': md5_string[27:29],
        'ngd4uy551': md5_string[17:19],
        'bd4uy742': md5_string[26:27],
        'ngd4yut78': md5_string[12:14],
        'iogojti': md5_string[25:26],
        'h67456y': md5_string[16:19],
        'lksytkjh': md5_string[17:21],
        'n3bf4uj7y7': md5_string[18:19],
        'nbf4uj7y432': md5_string[21:23],
        'ibvytiqjek': md5_string[14:16],
        'h13ey474': md5_string[29:32],
        'abiokytke': md5_string[21:23],
        'bd24y6421f': md5_string[24:26],
        'tbvdiuytk': md5_string[16:17],
    }

def fetch_pe_pb_xilv_data(gu_code, ts):
    """
    发起估值接口请求，返回 pe, pb, xilv 三个数值（float）。
    出错时返回 0,0,0。
    """
    try:
        t = f"{ts}{gu_code}pepcnew2.2.7-1EWf45rlv#kfsr@k#gfksgkr"
        md5_value = hashlib.md5(t.encode('utf-8')).hexdigest()
        md5_parts = split_md5(md5_value, ts, gu_code)
        body = json.dumps(md5_parts)
        headers2 = {
            "Host": "api.jiucaishuo.com",
            "Content-Type": "application/json;charset=UTF-8",
            "User-Agent": headers["User-Agent"],
        }
        r = requests.post("https://api.jiucaishuo.com/v2/guzhi/newtubiaodata", headers=headers2, data=body, timeout=10)
        if r.status_code == 200:
            data = r.json()
            # 取 new_percent_value 中的百分比数字并转 float
            pe_str = data.get('data', {}).get('top_data', [None, {}, {}, {}])[1].get('new_percent_value', {}).get('value', '0')
            pb_str = data.get('data', {}).get('top_data', [None, {}, {}, {}])[2].get('new_percent_value', {}).get('value', '0')
            xilv_str = data.get('data', {}).get('top_data', [None, {}, {}, {}])[3].get('new_percent_value', {}).get('value', '0')

            def parse_percent(s):
                try:
                    return float(str(s).replace('%', '').strip() or 0.0)
                except:
                    return 0.0

            pe = parse_percent(pe_str)
            pb = parse_percent(pb_str)
            xilv = parse_percent(xilv_str)
            # 保留两位小数
            pe = round(pe, 2)
            pb = round(pb, 2)
            xilv = round(xilv, 2)
            return pe, pb, xilv
    except Exception as e:
        print(f"估值接口出错: {e}")
    return 0.0, 0.0, 0.0

def update_pe_pb_xilv_to_ws(ws, target_col):
    for name, data in pe_pb_xilv.items():
        if data["row"] == 0:
            continue
        pe, pb, xilv = fetch_pe_pb_xilv_data(data["code"], int(time.time() * 1000))
        # 结果按 calc 权重计算，并保留两位小数
        try:
            result = pe * data["calc"][0] + pb * data["calc"][1] + xilv * data["calc"][2]
            result = round(float(result), 2)
        except Exception:
            result = 0.0
        write_number_cell(ws, data["row"], target_col, result)
        print(f"{name} 估值结果: \n\t代码: {data['code']}\n\tpe百分位: {pe} pb百分位: {pb} 息率: {xilv}\n\t权重: {data['calc']}\n\t结果: {result}")
        time.sleep(1)

# --------------------------
# 导出实时数据
# --------------------------
def export_realtime_data():
    xlsx_path = os.path.join(os.path.dirname(__file__), "stocks_data.xlsx")
    if not os.path.exists(xlsx_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "StockData"
    else:
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active

    last_col = detect_last_col(ws)
    target_col = last_col + 1

    fetch_stock_data_to_ws(ws, target_col)
    update_pe_pb_xilv_to_ws(ws, target_col)
    set_column_style(ws, target_col)

    wb.save(xlsx_path)

if __name__ == "__main__":
    export_realtime_data()
# End-903-2025.10.22.140534
