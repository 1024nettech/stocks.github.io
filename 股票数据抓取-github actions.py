import os
import requests
import time
import openpyxl
import json
import hashlib
import re
from datetime import datetime

# --------------------------
# Excel 工具函数
# --------------------------
def detect_last_col(ws):
    last_col = 1
    for col in range(1, ws.max_column + 1):
        column_values = [ws.cell(row=row, column=col).value for row in range(1, ws.max_row + 1)]
        if any(v is not None and v != "" for v in column_values):
            last_col = col
    return last_col

def set_column_style(ws, col_idx):
    for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.font = openpyxl.styles.Font(name="宋体", size=12)
            cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    ws.cell(row=2, column=col_idx).font = openpyxl.styles.Font(name="宋体", size=12, bold=True)

def safe_float_convert(val):
    try:
        return round(float(val), 2)
    except Exception:
        return val

def write_number_cell(ws, row, col, value):
    if isinstance(value, (int, float)):
        rounded = round(float(value), 2)
        ws.cell(row=row, column=col, value=rounded)
        ws.cell(row=row, column=col).number_format = '0.00'
    else:
        ws.cell(row=row, column=col, value=value)

# --------------------------
# 股票与指数配置
# --------------------------
stock_index_start_row = 7
stock_val_start_row = 6

stocks_index = {
    "上证点数": {
        "code": "000001",
        "row": 4,
        "result": "",
        "url": "https://qt.gtimg.cn/?q=s_sh",
    },
    "中证A500": {
        "code": "000510",
        "row": 0,
        "result": "",
        "url": "https://qt.gtimg.cn/?q=s_sh",
    },
}

pe_pb_xilv = {
    "沪深全A(万德全A)": {
        "row": 3,
        "code": "881001.WI",
        "calc": [0.5, 0.5, 0]
    },
    "中证A500": {
        "row": 0,
        "code": "000510.SH",
        "calc": [0.5, 0.5, 0]
    },
}

# 自动补充行号
row_value = stock_index_start_row
for v in stocks_index.values():
    if v["row"] == 0:
        v["row"] = row_value
        row_value += 3

row_value = stock_val_start_row
for v in pe_pb_xilv.values():
    if v["row"] == 0:
        v["row"] = row_value
        row_value += 3

# --------------------------
# 请求设置
# --------------------------
session = requests.Session()
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:145.0) Gecko/20100101 Firefox/145.0",
    "Accept": "*/*",
    "Accept-Language": "zh-CN",
    "Accept-Encoding": "gzip, deflate, br, zstd",
    "Sec-Fetch-Storage-Access": "none",
    "DNT": "1",
    "Sec-GPC": "1",
    "Connection": "keep-alive",
    "Referer": "https://finance.sina.com.cn/",
    "Sec-Fetch-Dest": "script",
    "Sec-Fetch-Mode": "no-cors",
    "Sec-Fetch-Site": "cross-site",
}

# --------------------------
# 抓取指数数据
# --------------------------
def fetch_stock_data_to_ws(ws, target_col):
    for name, data in stocks_index.items():
        url = data["url"] + data["code"]
        try:
            response = session.get(url, headers=headers, timeout=10)
            response.raise_for_status()
            parts = response.text.split("~")
            parsed = parts[3] if len(parts) > 3 and parts[3] else parts[0]
            val = safe_float_convert(parsed)
            data["result"] = val
            if data["row"] == 4:
                ws.cell(row=1, column=target_col, value=datetime.now().strftime("%Y/%m/%d"))
                ws.cell(row=2, column=target_col, value="上证")
            write_number_cell(ws, data["row"], target_col, val)
        except:
            pass
        time.sleep(1)

# --------------------------
# 获取 PE / PB / Xilv 数据
# --------------------------
def split_md5(md5_string, ts, gu_code):
    return {
        "gu_code": gu_code,
        "pe_category": "pe",
        "year": -1,
        "category": "",
        "ver": "new",
        "type": "pc",
        "version": "2.2.7",
        "authtoken": "",
        "act_time": ts,
    }

def fetch_pe_pb_xilv_data(gu_code, ts):
    try:
        t = f"{ts}{gu_code}pepcnew2.2.7-1EWf45rlv#kfsr@k#gfksgkr"
        md5_value = hashlib.md5(t.encode('utf-8')).hexdigest()
        body = json.dumps(split_md5(md5_value, ts, gu_code))
        headers2 = {
            "Host": "api.jiucaishuo.com",
            "Content-Type": "application/json;charset=UTF-8",
            "User-Agent": headers["User-Agent"],
        }
        r = requests.post("https://api.jiucaishuo.com/v2/guzhi/newtubiaodata", headers=headers2, data=body, timeout=10)
        data = r.json()
        pe_str = data.get('data', {}).get('top_data', [None, {}, {}, {}])[1].get('new_percent_value', {}).get('value', '0')
        pb_str = data.get('data', {}).get('top_data', [None, {}, {}, {}])[2].get('new_percent_value', {}).get('value', '0')
        xilv_str = data.get('data', {}).get('top_data', [None, {}, {}, {}])[3].get('new_percent_value', {}).get('value', '0')

        def parse_percent(s): return float(str(s).replace('%', '').strip() or 0.0)
        return round(parse_percent(pe_str), 2), round(parse_percent(pb_str), 2), round(parse_percent(xilv_str), 2)
    except:
        return 0.0, 0.0, 0.0

def update_pe_pb_xilv_to_ws(ws, target_col):
    for name, data in pe_pb_xilv.items():
        pe, pb, xilv = fetch_pe_pb_xilv_data(data["code"], int(time.time() * 1000))
        try:
            result = pe * data["calc"][0] + pb * data["calc"][1] + xilv * data["calc"][2]
            result = round(float(result), 2)
        except:
            result = 0.0
        write_number_cell(ws, data["row"], target_col, result)
        time.sleep(1)

# --------------------------
# 导出实时数据
# --------------------------
def export_realtime_data():
    xlsx_path = os.path.join(os.path.dirname(__file__), "stocks_data.xlsx")
    if not os.path.exists(xlsx_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "StockData"
    else:
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active

    last_col = detect_last_col(ws)
    target_col = last_col + 1

    fetch_stock_data_to_ws(ws, target_col)
    update_pe_pb_xilv_to_ws(ws, target_col)
    set_column_style(ws, target_col)

    wb.save(xlsx_path)

if __name__ == "__main__":
    export_realtime_data()
# End-201-2025.10.22.130023
